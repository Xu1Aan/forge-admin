from __future__ import annotations

import sys
from pathlib import Path


def main() -> int:
    p = Path(r"e:\project\forge-admin\data\项目表.xlsx")
    if not p.exists():
        print(f"not found: {p}")
        return 2

    try:
        import openpyxl  # type: ignore
    except Exception as e:
        print("openpyxl not available:", e)
        return 3

    wb = openpyxl.load_workbook(p, data_only=True)
    print("sheets:", wb.sheetnames)
    ws = wb[wb.sheetnames[0]]
    print("active_sheet:", ws.title)
    print("max_row:", ws.max_row, "max_col:", ws.max_column)

    def cell(r: int, c: int) -> str:
        v = ws.cell(r, c).value
        if v is None:
            return ""
        return str(v).strip()

    headers = [cell(1, c) for c in range(1, ws.max_column + 1)]
    print("headers:", headers)

    # show first few non-empty rows
    shown = 0
    for r in range(2, min(ws.max_row, 200) + 1):
        row = [cell(r, c) for c in range(1, ws.max_column + 1)]
        if not any(x for x in row):
            continue
        print(f"row{r}:", row)
        shown += 1
        if shown >= 10:
            break

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

